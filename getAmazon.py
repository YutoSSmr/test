from cgitb import text
from attr import attr
import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin
import openpyxl
from sqlalchemy import column


# ランキングの商品名、価格取得
def getRankInfo(res):
    rowNUm = 1
    rowNum2 = 2
    wb = openpyxl.load_workbook('amazonRankData.xlsx')
    ws = wb['Sheet1']
    #テキスト取得
    soup = BeautifulSoup(res.text,"html.parser")
    products = soup.find('div', attrs={'id':'zg_left_col1'})
    rankNum = len(products.find_all('div', attrs={'class':'celwidget'})) - 1
    print(rankNum)
    for i in range(rankNum):
        # ランキングカテゴリ取得
        category = products.find('div', attrs={'cel_widget_id':f'p13n-zg-list-carousel-desktop_zeitgeist-lists_{i}'})
        heading = category.find('h2')
        ws.cell(column=1, row=rowNUm).value = heading.text
        rowNUm += 3
        print(heading.text)
        for j in range(3):
            # 商品取得
            productsName = category.find_all('div', attrs={'class':'p13n-sc-truncate-desktop-type2'})
            ws.cell(column=j+1, row=rowNum2).value = productsName[j].text

            # 価格取得
            product = category.find_all('li')
            productPrice = product[j].find_all('span', attrs={'class':'a-size-base'})
            ws.cell(column=j+1, row=rowNum2+1).value = productPrice[0].text

        rowNum2 += 3    

    wb.save('amazonRankData.xlsx')

            
def main():

    # ランキングURLログイン
    ranking_url = "https://www.amazon.co.jp/ranking?type=top-sellers&ref_=nav_cs_bestsellers"
    res = requests.get(ranking_url)

    getRankInfo(res)



if __name__ == '__main__':
    main()

